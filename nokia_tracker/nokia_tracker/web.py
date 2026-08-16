"""Flask web UI — dashboard, portfel, dywidendy, newsy, prognozy, ustawienia
(krok 9, BLUEPRINT §3/§9). Cache-busting: no-store na HTML/API, statyki
?v=<wersja>, badge wersji w nav (CLAUDE.md — WebView Companion cache'uje
HTML agresywnie i nie rewaliduje).
"""
from __future__ import annotations

import csv
import io
import json
import logging
import os
import secrets
from datetime import date, datetime, timedelta, timezone
from pathlib import Path

import openpyxl
from flask import Flask, Response, redirect, render_template, request, url_for

from . import __version__, advisor as advisorm, analysis, backup as backupm, db as dbm, fx
from . import dashboard_insights
from . import format as fmt
from . import portfolio as portfoliom
from . import dividend_outlook as outlookm
from . import quotes, sensors
from . import settings as settingsm
from .analytics import attribution as analytics_attribution
from .analytics import benchmark as analytics_benchmark
from .analytics import returns as analytics_returns
from .analytics import risk as analytics_risk
from .importers import computershare_pdf
from .providers import fx_nbp
from .tax import dividends as taxdiv
from .tax import grants as grantsm
from .tax import losses as taxlosses
from .tax import lots as taxlots
from .tax import pit38 as taxpit38
from .tax import policy as taxpolicy
from .tax import trace as taxtrace
from .tax import whatif as taxwhatif
from .ai import chat as ai_chat
from .ai import copilot as ai_copilot
from .ai import openai_compat
from .ai import status as ai_status

logger = logging.getLogger(__name__)

_PRIMARY_SYMBOL = "NOKIA.HE"
_ERICSSON_SYMBOL = "ERIC-B.ST"
_OMXH25_SYMBOL = "^OMXH25"
_EURUSD_SYMBOL = "EURUSD=X"
_ADR_SYMBOL = "NOK"

# Zakresy wykresu pulpitu (krok 16): (granularity, dni_wstecz). `None` dni = bez
# filtra dolnego (całość dostępnej historii). "1d" jedyny na intraday — reszta na
# świecach dziennych, których backfill/refresh i tak już istnieje (quotes.py).
_CHART_RANGES: dict[str, tuple[str, int | None]] = {
    "1d": ("intraday", None),
    "1w": ("daily", 7),
    "1m": ("daily", 31),
    "3m": ("daily", 93),
    "6m": ("daily", 186),
    "1y": ("daily", 366),
    "3y": ("daily", 3 * 366),
    "5y": ("daily", 5 * 366),
    "max": ("daily", None),
}
_DEFAULT_CHART_RANGE = "3m"


def _ids(conn) -> dict:
    """Get-or-create instrumentów — niezależne od main.py (web.py może
    obsłużyć request zanim scheduler w main.py przejdzie backfill)."""
    return {
        "primary": quotes.ensure_instrument(conn, _PRIMARY_SYMBOL, "Nokia Oyj", "EUR", "primary"),
        "ericsson": quotes.ensure_instrument(conn, _ERICSSON_SYMBOL, "Ericsson", "SEK", "benchmark"),
        "omxh25": quotes.ensure_instrument(
            conn, _OMXH25_SYMBOL, "OMX Helsinki 25", "EUR", "benchmark"),
        "eurpln": quotes.ensure_instrument(conn, fx.EURPLN_SYMBOL, "EUR/PLN", "PLN", "fx"),
        "eurusd": quotes.ensure_instrument(conn, _EURUSD_SYMBOL, "EUR/USD", "USD", "fx"),
        "adr": quotes.ensure_instrument(conn, _ADR_SYMBOL, "Nokia ADR (NYSE)", "USD", "adr"),
    }


# Krok 18: kopia LOT_TYPE_LABELS z templates/_macros.html — ta strona żyje w
# Pythonie (JSON /api/preview/sale), tamta w Jinja (tabele HTML); nie da się
# ich zeszyć w jedno miejsce bez importowania Jinja env do zwykłej funkcji.
_LOT_TYPE_LABELS_PY = {"own": "własne", "matched": "podarowane", "lti": "LTI",
                       "dividend_drip": "dywidenda"}


def _is_future_date(date_str: str) -> bool:
    """Krok 16 (§8.2): NBP zwraca HTTP 400 dla dat przyszłych — bez tej
    walidacji `fx_nbp.rate_on_or_before` podnosi `QuoteProviderError`, który
    nie jest łapany w `tax/lots.py`/`tax/dividends.py`, więc formularz
    kończy się gołym 500 zamiast czytelnego komunikatu. Puste/niepoprawne
    daty NIE są tu odrzucane — to i tak zgłosi się inaczej (np. pusty
    `acquired_date`), walidujemy tylko to, co potrafimy jednoznacznie ocenić."""
    try:
        return date.fromisoformat(date_str) > date.today()
    except (TypeError, ValueError):
        return False


def _ai_keys() -> dict:
    """Klucze API z ENV — NIE z tabeli settings (patrz settings.py)."""
    return {
        "local_llm_api_key": os.environ.get("LOCAL_LLM_API_KEY", ""),
        "gemini_api_key": os.environ.get("GEMINI_API_KEY", ""),
        "anthropic_api_key": os.environ.get("ANTHROPIC_API_KEY", ""),
    }


class _IngressPrefixMiddleware:
    """HA ingress zdejmuje prefiks ścieżki przed przekazaniem requestu do
    kontenera (Flask widzi czyste '/', '/portfolio' itd.), ale oryginalny
    prefiks leci w nagłówku X-Ingress-Path. Bez ustawienia SCRIPT_NAME
    url_for()/redirect() generują URL-e bez prefiksu, które w przeglądarce
    rozwiązują się pod domeną główną HA zamiast pod ingressem (złapane na
    żywo Playwrightem: /static/app.css -> 404, bo poszło do
    homeassistant.local:8123/static/... zamiast pod ingress prefix)."""

    def __init__(self, wsgi_app):
        self.wsgi_app = wsgi_app

    def __call__(self, environ, start_response):
        prefix = environ.get("HTTP_X_INGRESS_PATH", "")
        if prefix:
            environ["SCRIPT_NAME"] = prefix
        return self.wsgi_app(environ, start_response)


def create_app(db_path: str) -> Flask:
    app = Flask(__name__)
    app.wsgi_app = _IngressPrefixMiddleware(app.wsgi_app)

    # Krok 23 (docs/PLAN_KROK_23_portfel_kafelki.md): formatowanie po polsku
    # (separator tysięcy, przecinek dziesiętny) dla karty „Portfel" na pulpicie.
    # Celowo NIE używane na stronach podatkowych (/lots, /pit38, ...) — te muszą
    # zostać bajtowo zestawialne z wyciągiem/NBP w dotychczasowym formacie.
    app.jinja_env.filters["money"] = fmt.money
    app.jinja_env.filters["qty"] = fmt.qty
    app.jinja_env.filters["pct"] = fmt.pct

    @app.after_request
    def _no_cache(resp: Response) -> Response:
        if resp.mimetype in ("text/html", "application/json"):
            resp.headers["Cache-Control"] = "no-store"
        return resp

    def _conn():
        return dbm.get_conn(db_path)

    @app.get("/")
    def dashboard():
        conn = _conn()
        try:
            ids = _ids(conn)
            values = sensors.market_values(conn, ids["primary"])
            values.update(sensors.benchmark_values(
                conn, ids["primary"], ids["ericsson"], ids["omxh25"], ids["eurpln"],
                ids["adr"], ids["eurusd"]))
            values.update(sensors.ai_values(conn))
            values.update(sensors.forecast_values(conn))

            cfg = settingsm.get_settings(conn)
            cost_basis_eur = cfg["position_qty"] * cfg["avg_cost_eur"]
            dividends = sensors.dividends_values(conn, cfg, cost_basis_eur)
            position = portfoliom.position_values_auto(
                conn, cfg, values.get("price_eur"), values.get("eurpln_rate"),
                dividends_net_total_eur=dividends["dividends_net_eur"])

            # Krok 21 (docs/PLAN_KROK_21_portfel_calkowity.md): zablokowane (nienabyte
            # dopasowania ESPP/transze LTI) i posiadane-ale-ograniczone (świeże zakupy
            # własne czekające na własne dopasowanie) — dotąd niewidoczne na pulpicie,
            # który pokazywał tylko "position" (same uwolnione akcje).
            unvested = grantsm.unvested_summary(
                conn, values.get("price_eur"), values.get("eurpln_rate"))
            restricted = grantsm.restricted_own_summary(
                conn, values.get("price_eur"), values.get("eurpln_rate"))
            # Krok 23 (docs/PLAN_KROK_23_portfel_kafelki.md): position/restricted/unvested
            # złożone w trzy kubełki (wolne/z ograniczeniem/zablokowane) + sumę — zastępuje
            # ręczną arytmetykę total_qty/total_value_* dawniej inline tutaj.
            buckets = portfoliom.dashboard_buckets(position, restricted, unvested)
            # Krok 26 (docs/PLAN_KROK_26_doradca.md): kwota przepadującego dopasowania
            # ESPP dopisana do istniejącego zdania ostrzegawczego (dawniej bez kwoty).
            forfeit = advisorm.forfeit_summary(
                conn, values.get("price_eur"), values.get("eurpln_rate"))

            # Krok 18: metadane kursu EUR/PLN (skąd, kiedy) dla linii rozgraniczającej
            # kurs bieżący (Yahoo/ECB, prezentacyjny) od kursu NBP zamrożonego na
            # zdarzenie (podatkowy, patrz /dywidendy i /pit38). Sama wartość kursu
            # to już `values["eurpln_rate"]` (sensors.py::benchmark_values).
            eurpln_row = quotes.latest_quote(conn, ids["eurpln"], granularity="daily")
            fx_info = {
                "rate": eurpln_row["close"] if eurpln_row else None,
                "ts": eurpln_row["ts"] if eurpln_row else None,
                "source": eurpln_row["source"] if eurpln_row else None,
            }

            recent_alerts = conn.execute(
                "SELECT * FROM alerts_log ORDER BY fired_at DESC LIMIT 5").fetchall()

            # Krok 28.5 (docs/PLAN_KROK_28_ux_mobile.md §5): "Dziś warto wiedzieć" —
            # reużywa dane już policzone powyżej (unvested) + dwa lekkie odczyty
            # (dostępna strata, dochód tego roku wg aktywnej polityki), zero nowej
            # matematyki podatkowej — te same funkcje co /pit38 i /pit38/kreator.
            current_year = datetime.now().year
            loss_available_pln = taxlosses.available_for_year(
                conn, cfg, current_year)["total_remaining_pln"]
            income_pln_this_year = taxpolicy.compute_all_policies(
                conn, cfg, current_year)[cfg["cost_basis_policy"]]["income_pln"]
            insights = dashboard_insights.today_worth_knowing(
                change_pct_day=values.get("change_pct_day"),
                next_vest_date=unvested.get("next_vest_date"),
                next_vest_qty=unvested.get("next_vest_qty"),
                loss_available_pln=loss_available_pln,
                income_pln_this_year=income_pln_this_year)

            return render_template(
                "dashboard.html", active="dashboard", version=__version__,
                values=values, position=position, dividends=dividends, fx_info=fx_info,
                unvested=unvested, restricted=restricted, buckets=buckets, forfeit=forfeit,
                chart_ranges=list(_CHART_RANGES), default_chart_range=_DEFAULT_CHART_RANGE,
                chart_api_url=url_for("chart_api"),
                alerts=[dict(r) for r in recent_alerts],
                insights=insights,
            )
        finally:
            conn.close()

    @app.get("/api/chart")
    def chart_api():
        """Krok 16: zasila konfigurowalny wykres pulpitu — `?range=` z
        `_CHART_RANGES` (1d na intraday, reszta na dziennych). `no-store`
        łapie się automatycznie (`_no_cache` obsługuje `application/json`),
        więc WebView Companion nie zserwuje kiedyś złapanego zakresu na
        stałe (patrz CLAUDE.md — cache mobilny)."""
        conn = _conn()
        try:
            range_key = request.args.get("range", _DEFAULT_CHART_RANGE)
            granularity, days = _CHART_RANGES.get(range_key, _CHART_RANGES[_DEFAULT_CHART_RANGE])
            since = (
                (datetime.now(timezone.utc) - timedelta(days=days)).isoformat()
                if days is not None else None)
            ids = _ids(conn)
            points = quotes.closes_in_range(conn, ids["primary"], granularity, since)
            return {"range": range_key, "granularity": granularity, "points": points}
        finally:
            conn.close()

    @app.post("/analyze-now")
    def analyze_now():
        conn = _conn()
        try:
            with dbm.WRITE_LOCK:
                ids = _ids(conn)
                cfg = dict(settingsm.get_settings(conn), **_ai_keys())
                ok = analysis.run_daily_analysis(
                    conn, cfg, ids["primary"], ids["ericsson"], ids["omxh25"], ids["eurpln"])
            return redirect(url_for("dashboard", analyzed="1" if ok else "0"))
        finally:
            conn.close()

    @app.get("/portfolio")
    def portfolio_get():
        conn = _conn()
        try:
            cfg = settingsm.get_settings(conn)
            lots_position = None
            if taxlots.open_lots(conn):
                ids = _ids(conn)
                price = quotes.latest_quote(conn, ids["primary"], granularity="daily")
                eurpln = quotes.latest_quote(conn, ids["eurpln"], granularity="daily")
                lots_position = portfoliom.lots_based_position_values(
                    conn, cfg, price["close"] if price else None,
                    eurpln["close"] if eurpln else None)
            return render_template(
                "portfolio.html", active="portfolio", version=__version__, cfg=cfg,
                lots_position=lots_position,
                saved=request.args.get("saved") == "1")
        finally:
            conn.close()

    @app.post("/portfolio")
    def portfolio_post():
        conn = _conn()
        try:
            qty = float(request.form.get("position_qty") or 0)
            avg_cost = float(request.form.get("avg_cost_eur") or 0)
            with dbm.WRITE_LOCK:
                settingsm.set_settings(conn, {"position_qty": qty, "avg_cost_eur": avg_cost})
            return redirect(url_for("portfolio_get", saved="1"))
        finally:
            conn.close()

    @app.get("/dividends")
    def dividends_get():
        """Krok 16: JEDNO źródło prawdy z `add_dividend()` — kwoty w PLN na kursie
        NBP zamrożonym na Record Date (`compute_dividend_tax_pln`, ten sam
        mechanizm co `/pit38`), nie osobny kalkulator EUR na bieżących stawkach
        jak przed ujednoliceniem formularza. `backfill_missing_dividend_rates`
        dogania dywidendy wpisane ręcznie przed tym krokiem (surowy INSERT
        wtedy nie zamrażał kursu).

        Krok 18: `totals` liczone SUMOWANIEM `items` (a nie osobnym wywołaniem
        `sensors.dividends_values`) — przed tą zmianą strona pokazywała dwie
        niezgodne matematyki 40px od siebie: kafelki na kursach BIEŻĄCYCH w EUR
        (`sensors.dividends_values`), tabela pod nimi na kursach NBP ZAMROŻONYCH
        na Record Date w PLN (`compute_dividend_tax_pln`, ten sam co tu). Zero
        nowych zapytań do NBP — `items` już ma policzone `*_pln` per wiersz.
        `sensors.dividends_values` zostaje nietknięte dla sensorów MQTT i dla
        linii dywidend na pulpicie (tam liczone po kursie bieżącym, spójnie z
        resztą pulpitu — patrz krok 2)."""
        conn = _conn()
        try:
            cfg = settingsm.get_settings(conn)
            taxdiv.backfill_missing_dividend_rates(conn)
            rows = conn.execute("SELECT * FROM dividends ORDER BY pay_date DESC").fetchall()
            items = []
            for r in rows:
                t = taxdiv.compute_dividend_tax_pln(r, cfg)
                d = dict(r)
                d.update(t)
                if d.get("reinvested_lot_id"):
                    lot = conn.execute(
                        "SELECT acquired_date, quantity, price_eur FROM lots WHERE id = ?",
                        (d["reinvested_lot_id"],)).fetchone()
                    d["reinvested_lot"] = dict(lot) if lot else None
                items.append(d)

            gross_pln = sum(i["gross_pln"] or 0 for i in items)
            withholding_paid_pln = sum(i.get("withholding_paid_pln") or 0 for i in items)

            # Krok 18: kafelki EUR pod PLN — NIE jest to osobna konwersja walutowa (a
            # więc nie wraca "dwóch matematyk"): `compute_dividend_tax` liczy tymi samymi
            # % (u źródła z wiersza / traktat / Belka z `cfg`) co `compute_dividend_tax_pln`
            # powyżej, po prostu bez mnożenia przez zamrożony kurs NBP — EUR to naturalna
            # waluta wypłaty, PLN to waluta rozliczenia z fiskusem. Jedno źródło procentów.
            eur_totals = [
                taxdiv.compute_dividend_tax(
                    i["gross_eur"],
                    i["withholding_pct"] if i["withholding_pct"] is not None
                    else cfg["finnish_withholding_pct"],
                    cfg["treaty_withholding_pct"], cfg["pl_capital_gains_tax_pct"])
                for i in items
            ]
            gross_eur = sum(i["gross_eur"] for i in items)

            # Krok 18: dawniej `cost_basis_eur = cfg["position_qty"] * cfg["avg_cost_eur"]` —
            # pola ręczne z ustawień, które po pierwszym imporcie PDF są zerami (stan
            # posiadania żyje w lotach). Kafelek „Yield on cost" pokazywał wtedy `—` na
            # stałe u każdego użytkownika po imporcie. `position_values_auto` przełącza
            # się na loty automatycznie, tak jak już robi to pulpit (web.py:141-143).
            position = portfoliom.position_values_auto(conn, cfg, None, None)
            cost_basis_eur = position["cost_basis_eur"]
            totals = {
                "dividends_gross_pln": gross_pln,
                "dividends_gross_eur": gross_eur,
                "withholding_paid_pln": withholding_paid_pln,
                "withholding_paid_eur": sum(t["withholding_paid_eur"] for t in eur_totals),
                "dividends_net_pln": gross_pln - withholding_paid_pln,
                "dividends_net_eur": sum(t["net_received_eur"] for t in eur_totals),
                "pl_tax_due_pln": sum(i.get("pl_tax_due_pln") or 0 for i in items),
                "pl_tax_due_eur": sum(t["pl_tax_due_eur"] for t in eur_totals),
                "reclaimable_from_finland_pln": sum(
                    i.get("reclaimable_from_finland_pln") or 0 for i in items),
                "reclaimable_from_finland_eur": sum(
                    t["reclaimable_from_finland_eur"] for t in eur_totals),
                "dividend_yield_on_cost_pct": (
                    gross_eur / cost_basis_eur * 100 if cost_basis_eur else None),
            }
            # Krok 28.4 (docs/PLAN_KROK_28_ux_mobile.md §4): słupki dywidend rok po
            # roku — agregacja `gross_pln`/`net_pln` już policzonych powyżej per
            # wiersz, zero nowej logiki podatkowej.
            yearly: dict[str, dict] = {}
            for i in items:
                y = i["pay_date"][:4]
                bucket = yearly.setdefault(y, {"gross_pln": 0.0, "net_pln": 0.0})
                bucket["gross_pln"] += i["gross_pln"] or 0
                bucket["net_pln"] += (i["gross_pln"] or 0) - (i.get("withholding_paid_pln") or 0)
            yearly_dividends = [
                {"year": y, "gross_pln": round(v["gross_pln"], 2), "net_pln": round(v["net_pln"], 2)}
                for y, v in sorted(yearly.items())]

            # Krok 30 (docs/PLAN_KROK_30_dywidendy.md): kalendarz i prognoza dywidend.
            # `reconcile_schedule` dotyka tylko rat jeszcze niedopasowanych (indeks na
            # `record_date`, tania operacja) — pod WRITE_LOCK, mimo że
            # `backfill_missing_dividend_rates` wyżej nie jest (przedkrokowy stan, nie
            # naprawiany tutaj, ale nowy zapis dostaje właściwy kontrakt od razu).
            with dbm.WRITE_LOCK:
                outlookm.reconcile_schedule(conn)
            schedule = outlookm.list_schedule(conn)
            lata_raw = request.args.get("lata")
            years_ahead = int(lata_raw) if lata_raw in ("1", "3", "5") else 3
            ids = _ids(conn)
            eurpln_row = quotes.latest_quote(conn, ids["eurpln"], granularity="daily")
            eurpln_rate = eurpln_row["close"] if eurpln_row else None
            outlook = outlookm.calendar(conn, cfg, years_ahead=years_ahead,
                                        eurpln_rate=eurpln_rate)

            return render_template(
                "dividends.html", active="dividends", version=__version__,
                items=items, totals=totals, cfg=cfg, saved=request.args.get("saved") == "1",
                error=request.args.get("error"), yearly_dividends=yearly_dividends,
                schedule=schedule, outlook=outlook, years_ahead=years_ahead)
        finally:
            conn.close()

    @app.post("/dividends/harmonogram")
    def dividend_schedule_post():
        """Krok 30: jedno ogłoszenie WZA = jeden formularz, do 4 rat naraz. Puste raty
        (pola bez `record_date`/`per_share`) są pomijane, nie zapisywane jako zera —
        WZA nie zawsze uchwala od razu wszystkie 4 daty. Świadomie BEZ
        `_is_future_date` — daty przyszłe są całym sensem tej tabeli (harmonogram
        dotyczy wypłat, które jeszcze się nie odbyły), a `dividend_schedule` nigdy
        nie dotyka NBP, więc walidacja stworzona dla `/lots`/`/dividends` tu by tylko
        po cichu wyłączyła funkcję."""
        conn = _conn()
        try:
            fiscal_year_raw = request.form.get("fiscal_year")
            try:
                fiscal_year = int(fiscal_year_raw)
            except (TypeError, ValueError):
                return redirect(url_for(
                    "dividends_get", error="Podaj rok obrotowy harmonogramu"))
            announced_on = request.form.get("announced_on") or None

            saved_any = False
            with dbm.WRITE_LOCK:
                for instalment in range(1, 5):
                    record_date = request.form.get(f"record_date_{instalment}")
                    per_share_raw = request.form.get(f"per_share_{instalment}")
                    if not record_date or not per_share_raw:
                        continue
                    payment_date = request.form.get(f"payment_date_{instalment}") or None
                    confirmed = bool(request.form.get(f"confirmed_{instalment}"))
                    outlookm.add_instalment(
                        conn, fiscal_year=fiscal_year, instalment=instalment,
                        record_date=record_date, gross_per_share_eur=float(per_share_raw),
                        payment_date=payment_date, dates_confirmed=confirmed,
                        announced_on=announced_on)
                    saved_any = True

            if not saved_any:
                return redirect(url_for(
                    "dividends_get",
                    error="Wypełnij co najmniej jedną ratę harmonogramu (data + stawka)"))
            return redirect(url_for("dividends_get", saved="1"))
        finally:
            conn.close()

    @app.post("/dividends/harmonogram/<int:schedule_id>/delete")
    def dividend_schedule_delete(schedule_id: int):
        conn = _conn()
        try:
            with dbm.WRITE_LOCK:
                outlookm.delete_instalment(conn, schedule_id)
            return redirect(url_for("dividends_get", saved="1"))
        finally:
            conn.close()

    @app.post("/dividends")
    def dividends_post():
        """Krok 16: przechodzi przez `taxdiv.add_dividend()` — jedyne miejsce
        zapisu dywidend (import PDF i formularz ręczny razem), więc kurs NBP
        zamrożony na Record Date i (opcjonalny) lot DRIP powstają identycznie
        niezależnie od źródła wpisu. Formularz nadal przyjmuje procent u
        źródła (nie kwotę), więc przeliczamy go na `taxes_eur` przed
        wywołaniem — `add_dividend` sam odtworzy ten sam procent z
        `taxes_eur/gross_eur`."""
        conn = _conn()
        try:
            pay_date = request.form.get("pay_date") or ""
            if _is_future_date(pay_date):
                return redirect(url_for(
                    "dividends_get", error="Data wypłaty nie może być w przyszłości "
                                           "(NBP nie publikuje kursów na przyszłe daty)"))
            drip_purchase_date = request.form.get("drip_purchase_date") or None
            if drip_purchase_date and _is_future_date(drip_purchase_date):
                return redirect(url_for(
                    "dividends_get", error="Data reinwestycji nie może być w przyszłości"))

            cfg = settingsm.get_settings(conn)
            gross_eur = float(request.form.get("gross_eur") or 0)
            quantity = float(request.form.get("quantity") or 0) or None
            gross_per_share = float(request.form.get("gross_per_share_eur") or 0) or None
            withholding_raw = request.form.get("withholding_pct")
            withholding_pct = (float(withholding_raw) if withholding_raw
                               else cfg["finnish_withholding_pct"])
            taxes_eur = gross_eur * withholding_pct / 100

            drip_price_raw = request.form.get("drip_price_eur")
            drip_shares_raw = request.form.get("drip_shares")
            purchase_price_eur = float(drip_price_raw) if drip_price_raw else None
            purchased_shares = float(drip_shares_raw) if drip_shares_raw else None

            # Klucz deterministyczny na treści formularza (nie na czasie zapisu):
            # przypadkowy podwójny submit tego samego wpisu jest teraz idempotentny
            # (poprawa względem starego surowego INSERT-a, który dublował wiersz).
            natural_key = f"manual:{pay_date}:{gross_eur}:{quantity or 0}:{withholding_pct}"

            with dbm.WRITE_LOCK:
                taxdiv.add_dividend(
                    conn, record_date=pay_date, entitled_quantity=quantity or 0.0,
                    gross_eur=gross_eur, taxes_eur=taxes_eur,
                    gross_per_share_eur=gross_per_share,
                    purchase_date=drip_purchase_date, purchase_price_eur=purchase_price_eur,
                    purchased_shares=purchased_shares, natural_key=natural_key)
            return redirect(url_for("dividends_get", saved="1"))
        finally:
            conn.close()

    @app.get("/lots")
    def lots_get():
        conn = _conn()
        try:
            cfg = settingsm.get_settings(conn)
            taxlots.backfill_missing_rates(conn)
            rows = conn.execute(
                "SELECT * FROM lots ORDER BY acquired_date DESC, id DESC").fetchall()
            year = cfg.get("tax_year") or None
            policies = taxpolicy.compute_all_policies(conn, cfg, year=year)
            return render_template(
                "lots.html", active="lots", version=__version__,
                lots=[dict(r) for r in rows], policies=policies, cfg=cfg,
                saved=request.args.get("saved") == "1",
                sold=request.args.get("sold") == "1",
                error=request.args.get("error"))
        finally:
            conn.close()

    @app.post("/lots")
    def lots_post():
        conn = _conn()
        try:
            acquired_date = request.form.get("acquired_date") or ""
            if _is_future_date(acquired_date):
                return redirect(url_for(
                    "lots_get", error="Data nabycia nie może być w przyszłości "
                                      "(NBP nie publikuje kursów na przyszłe daty)"))
            lot_type = request.form.get("lot_type") or "own"
            quantity = float(request.form.get("quantity") or 0)
            price_eur = float(request.form.get("price_eur") or 0)
            fee_eur = float(request.form.get("fee_eur") or 0)
            with dbm.WRITE_LOCK:
                taxlots.add_lot(conn, acquired_date, lot_type, quantity, price_eur,
                                fee_eur=fee_eur)
            return redirect(url_for("lots_get", saved="1"))
        finally:
            conn.close()

    @app.post("/lots/sell")
    def lots_sell_post():
        conn = _conn()
        try:
            sale_date = request.form.get("sale_date") or ""
            if _is_future_date(sale_date):
                return redirect(url_for(
                    "lots_get", error="Data sprzedaży nie może być w przyszłości "
                                      "(NBP nie publikuje kursów na przyszłe daty)"))
            quantity = float(request.form.get("sale_quantity") or 0)
            price_eur = float(request.form.get("sale_price_eur") or 0)
            fee_eur = float(request.form.get("sale_fee_eur") or 0)
            proceeds_raw = request.form.get("sale_proceeds_eur") or ""
            proceeds_eur = float(proceeds_raw) if proceeds_raw.strip() else None
            try:
                with dbm.WRITE_LOCK:
                    taxlots.record_sale(
                        conn, sale_date, quantity, price_eur, fee_eur=fee_eur,
                        proceeds_eur=proceeds_eur)
                return redirect(url_for("lots_get", sold="1"))
            except (taxlots.InsufficientLotsError, taxlots.CostBasisMissingError) as e:
                return redirect(url_for("lots_get", error=str(e)))
        finally:
            conn.close()

    # Krok 18: podgląd na żywo przed zapisem — dywidenda/lot/sprzedaż. Trzy
    # zasady dla wszystkich trzech endpointów: (1) tylko odczyt, zero zapisu do
    # `lots`/`sales`/`dividends`/`sale_allocations` (2) ten sam silnik co POST,
    # żeby podgląd nigdy nie rozjechał się z tym, co realnie zostanie zapisane
    # (3) błędy (brak pokrycia, brak kursu NBP, data w przyszłości) wracają jako
    # `{ok: false, error: ...}` z HTTP 200, nigdy 500 — formularz ma ostrzegać,
    # nie się wywalać. `rate_for_event`/`simulate_sale` mogą zrobić `INSERT OR
    # IGNORE` do `nbp_rates` (cache kursów publicznych) — nie do encji domenowych.

    @app.get("/api/preview/lot")
    def preview_lot():
        conn = _conn()
        try:
            acquired_date = request.args.get("acquired_date") or ""
            if not acquired_date:
                return {"ok": False, "error": "Podaj datę nabycia."}
            if _is_future_date(acquired_date):
                return {"ok": False, "error": "Data nabycia nie może być w przyszłości "
                                              "(NBP nie publikuje kursów na przyszłe daty)."}
            try:
                quantity = float(request.args.get("quantity") or 0)
                price_eur = float(request.args.get("price_eur") or 0)
                fee_eur = float(request.args.get("fee_eur") or 0)
            except ValueError:
                return {"ok": False, "error": "Niepoprawna liczba."}
            if quantity <= 0 or price_eur <= 0:
                return {"ok": False, "error": "Podaj ilość i cenę większe od zera."}

            rate = fx_nbp.rate_for_event(conn, acquired_date)
            if rate is None:
                return {"ok": False,
                        "error": f"Brak kursu NBP dla dnia {acquired_date} "
                                 "(spróbuj ponownie później)."}
            nbp_rate, nbp_rate_date = rate
            cost_eur = quantity * price_eur + fee_eur
            cost_pln = cost_eur * nbp_rate
            deriv = taxtrace.fx_derivation(conn, acquired_date, nbp_rate, nbp_rate_date, "nabycie")
            return {
                "ok": True,
                "nbp_rate": nbp_rate,
                "nbp_rate_date": nbp_rate_date,
                "explanation_pl": deriv["explanation_pl"],
                "table_urls": deriv.get("urls"),
                "lines": [
                    {"label": "Koszt", "value": round(cost_eur, 2), "unit": "EUR"},
                    {"label": "Koszt", "value": round(cost_pln, 2), "unit": "PLN",
                     "emphasis": True},
                ],
            }
        finally:
            conn.close()

    @app.get("/api/preview/sale")
    def preview_sale():
        conn = _conn()
        try:
            # sale_date opcjonalna — /pit38 „co jeśli sprzedam teraz" nie zbiera
            # daty (zawsze dziś, tak samo jak `simulate_sale(sale_date=None)`);
            # /lots „Zarejestruj sprzedaż" ją wysyła i wtedy jest walidowana.
            sale_date = request.args.get("sale_date") or None
            if sale_date and _is_future_date(sale_date):
                return {"ok": False, "error": "Data sprzedaży nie może być w przyszłości "
                                              "(NBP nie publikuje kursów na przyszłe daty)."}
            try:
                quantity = float(request.args.get("quantity") or 0)
                price_eur = float(request.args.get("price_eur") or 0)
                fee_eur = float(request.args.get("fee_eur") or 0)
            except ValueError:
                return {"ok": False, "error": "Niepoprawna liczba."}
            if quantity <= 0 or price_eur <= 0:
                return {"ok": False, "error": "Podaj ilość i cenę większe od zera."}

            cfg = settingsm.get_settings(conn)
            try:
                result = taxwhatif.simulate_sale(conn, cfg, quantity, price_eur, fee_eur, sale_date)
            except (taxlots.InsufficientLotsError, taxlots.CostBasisMissingError) as e:
                return {"ok": False, "error": str(e)}

            active = result["policies"][result["active_policy"]]
            sale_fx = (result["lots_consumed_detailed"] or {}).get("sale_fx") or {}
            lots_line = " · ".join(
                f"{_LOT_TYPE_LABELS_PY.get(a['lot_type'], a['lot_type'])} "
                f"{a['quantity']:.4f} @ {a['acquired_date']}"
                for a in result["lots_consumed"])
            return {
                "ok": True,
                "nbp_rate": result["nbp_rate"],
                "nbp_rate_date": result["nbp_rate_date"],
                "explanation_pl": sale_fx.get("explanation_pl"),
                "table_urls": sale_fx.get("urls"),
                "lines": [
                    {"label": "Loty (FIFO)", "value": lots_line or "—", "unit": None},
                    {"label": "Przychód", "value": result["revenue_pln"], "unit": "PLN"},
                    {"label": "Koszt", "value": active["cost_pln"], "unit": "PLN"},
                    {"label": "Podatek", "value": active["tax_pln"], "unit": "PLN"},
                    {"label": "Na rękę", "value": result["net_proceeds_pln"], "unit": "PLN",
                     "emphasis": True},
                ],
            }
        finally:
            conn.close()

    @app.get("/api/preview/dividend")
    def preview_dividend():
        conn = _conn()
        try:
            pay_date = request.args.get("pay_date") or ""
            if not pay_date:
                return {"ok": False, "error": "Podaj datę wypłaty."}
            if _is_future_date(pay_date):
                return {"ok": False, "error": "Data wypłaty nie może być w przyszłości "
                                              "(NBP nie publikuje kursów na przyszłe daty)."}
            try:
                gross_eur = float(request.args.get("gross_eur") or 0)
            except ValueError:
                return {"ok": False, "error": "Niepoprawna liczba."}
            if gross_eur <= 0:
                return {"ok": False, "error": "Podaj kwotę brutto większą od zera."}

            cfg = settingsm.get_settings(conn)
            withholding_raw = request.args.get("withholding_pct")
            withholding_pct = (float(withholding_raw) if withholding_raw
                               else cfg["finnish_withholding_pct"])

            rate = fx_nbp.rate_for_event(conn, pay_date)
            if rate is None:
                return {"ok": False,
                        "error": f"Brak kursu NBP dla dnia {pay_date} (spróbuj ponownie później)."}
            nbp_rate, nbp_rate_date = rate
            gross_pln = gross_eur * nbp_rate
            tax = taxdiv.compute_dividend_tax_pln(
                {"gross_pln": gross_pln, "withholding_pct": withholding_pct}, cfg)
            deriv = taxtrace.fx_derivation(conn, pay_date, nbp_rate, nbp_rate_date, "dywidenda")

            lines = [
                {"label": "Brutto", "value": round(gross_pln, 2), "unit": "PLN"},
                {"label": "Pobrane u źródła", "value": tax["withholding_paid_pln"], "unit": "PLN"},
                {"label": "Belka (19%)", "value": tax["belka_pln"], "unit": "PLN"},
                {"label": "Dopłata w PL", "value": tax["pl_tax_due_pln"], "unit": "PLN",
                 "emphasis": True},
                {"label": "Do odzyskania z Vero", "value": tax["reclaimable_from_finland_pln"],
                 "unit": "PLN"},
            ]

            drip_shares_raw = request.args.get("drip_shares")
            drip_price_raw = request.args.get("drip_price_eur")
            drip_date = request.args.get("drip_purchase_date")
            if drip_shares_raw and drip_price_raw and drip_date:
                try:
                    drip_shares = float(drip_shares_raw)
                    drip_price = float(drip_price_raw)
                    lines.append({
                        "label": "Powstanie lot",
                        "value": f"{drip_shares:.4f} akcji @ {drip_price:.4f} EUR ({drip_date})",
                        "unit": None,
                    })
                except ValueError:
                    pass

            return {
                "ok": True,
                "nbp_rate": nbp_rate,
                "nbp_rate_date": nbp_rate_date,
                "explanation_pl": deriv["explanation_pl"],
                "table_urls": deriv.get("urls"),
                "lines": lines,
            }
        finally:
            conn.close()

    @app.get("/sales")
    def sales_get():
        """Zrealizowane sprzedaże — pełne rozbicie do numeru tabeli NBP per
        sprzedaż (krok 16), tym samym `_alloc_detail.html`/`tax/trace.py` co
        karta „co jeśli sprzedam teraz" na `/pit38` — jedno źródło matematyki
        i formatowania dla symulacji i rzeczywistości."""
        conn = _conn()
        try:
            cfg = settingsm.get_settings(conn)
            year = request.args.get("year", type=int)
            query = "SELECT * FROM sales"
            params: tuple = ()
            if year:
                query += " WHERE strftime('%Y', sale_date) = ?"
                params = (str(year),)
            query += " ORDER BY sale_date DESC, id DESC"
            sale_rows = conn.execute(query, params).fetchall()

            sales = []
            for s in sale_rows:
                allocations = [dict(r) for r in conn.execute(
                    "SELECT lot_id, quantity, cost_pln, revenue_pln FROM sale_allocations "
                    "WHERE sale_id = ? ORDER BY lot_id", (s["id"],)).fetchall()]
                sale_ctx = {
                    "sale_date": s["sale_date"], "price_eur": s["price_eur"],
                    "fee_eur": s["fee_eur"], "quantity": s["quantity"],
                    "nbp_rate": s["nbp_rate"], "nbp_rate_date": s["nbp_rate_date"],
                }
                reported = None
                if s["reported_revenue_pln"] is not None or s["reported_cost_pln"] is not None:
                    reported = {"reported_revenue_pln": s["reported_revenue_pln"],
                                "reported_cost_pln": s["reported_cost_pln"]}
                detail = taxtrace.enrich_allocations(conn, allocations, sale_ctx, cfg, reported)
                sales.append({"sale": dict(s), "detail": detail})

            # Krok 17: pasek KPI nad rejestrem — suma wg AKTYWNEJ polityki kosztu,
            # ta sama, którą pokazuje szczegół każdej sprzedaży.
            active_policy = cfg.get("cost_basis_policy", "own_only")
            totals = {
                "count": len(sales),
                "revenue_pln": round(sum(i["detail"]["revenue_pln"] for i in sales), 2),
                "cost_pln": round(sum(
                    i["detail"]["policies"][active_policy]["cost_pln"] for i in sales), 2),
                "income_pln": round(sum(
                    i["detail"]["policies"][active_policy]["income_pln"] for i in sales), 2),
                "tax_pln": round(sum(
                    i["detail"]["policies"][active_policy]["tax_pln"] for i in sales), 2),
                "net_pln": round(sum(i["detail"]["net_pln"] for i in sales), 2),
            }

            return render_template(
                "sales.html", active="sales", version=__version__, cfg=cfg,
                sales=sales, year=year, totals=totals,
                deleted=request.args.get("deleted") == "1")
        finally:
            conn.close()

    @app.post("/sales/<int:sale_id>/delete")
    def sales_delete(sale_id: int):
        conn = _conn()
        try:
            with dbm.WRITE_LOCK:
                taxlots.reverse_sale(conn, sale_id)
            return redirect(url_for("sales_get", deleted="1"))
        finally:
            conn.close()

    @app.post("/sales/<int:sale_id>/report")
    def sales_report(sale_id: int):
        """Krok 20: zgłoszona wartość sprzedaży (np. zgodnie z ręcznym arkuszem
        użytkownika, gdy deklaracja już złożona i świadomie NIE jest korygowana —
        patrz docs/PLAN_KROK_20_reported_override.md). Nadpisuje TYLKO agregat
        PIT-38 (`tax/policy.py::compute_all_policies`) — `sale_allocations`/`lots`
        (realny ślad FIFO) zostają nietknięte. Puste pole = usuń nadpisanie
        (wróć do wyliczenia silnika)."""
        conn = _conn()
        try:
            exists = conn.execute("SELECT 1 FROM sales WHERE id = ?", (sale_id,)).fetchone()
            if not exists:
                return redirect(url_for("sales_get"))
            revenue_raw = (request.form.get("reported_revenue_pln") or "").strip()
            cost_raw = (request.form.get("reported_cost_pln") or "").strip()
            note_raw = (request.form.get("reported_note") or "").strip() or None
            reported_revenue = float(revenue_raw) if revenue_raw else None
            reported_cost = float(cost_raw) if cost_raw else None
            with dbm.WRITE_LOCK:
                conn.execute(
                    "UPDATE sales SET reported_revenue_pln = ?, reported_cost_pln = ?, "
                    "notes = ? WHERE id = ?",
                    (reported_revenue, reported_cost, note_raw, sale_id))
                conn.commit()
            return redirect(url_for("sales_get", reported="1"))
        finally:
            conn.close()

    @app.get("/grants")
    def grants_get():
        """Krok 16: dociąga bieżącą cenę/kurs dokładnie tak jak `portfolio_get`
        (patrz `lots_based_position_values` wyżej) i dokłada wycenę per transza
        (`tax/grants.py::valuation`) — aktualną dla części otwartej, z dnia
        sprzedaży dla części zrealizowanej."""
        conn = _conn()
        try:
            ids = _ids(conn)
            price = quotes.latest_quote(conn, ids["primary"], granularity="daily")
            eurpln = quotes.latest_quote(conn, ids["eurpln"], granularity="daily")
            valuation = grantsm.valuation(
                conn, price["close"] if price else None, eurpln["close"] if eurpln else None)

            espp = grantsm.list_espp(conn)
            lti = grantsm.list_lti_grouped(conn)
            # Krok 18: `sensors.grants_values` już liczy to dla MQTT — strona *o
            # vestingu* go dotąd nie pokazywała wcale.
            vesting = sensors.grants_values(conn)
            return render_template(
                "grants.html", active="grants", version=__version__,
                espp=espp, lti=lti, valuation=valuation, vesting=vesting)
        finally:
            conn.close()

    @app.get("/wyniki")
    def wyniki_get():
        """Krok 25 (docs/PLAN_KROK_25_wyniki.md): XIRR na wpłatach własnych,
        TWR z materializowanej `portfolio_history` (przeliczanej nocnym jobem
        — `rebuild_portfolio_history_job` w main.py), atrybucja zysku,
        kontrfaktyczny benchmark OMXH25 — jako krzywa (`counterfactual_series`)
        obok krzywej wartości portfela na tym samym wykresie."""
        conn = _conn()
        try:
            ids = _ids(conn)
            cfg = settingsm.get_settings(conn)
            price_row = quotes.latest_quote(conn, ids["primary"], granularity="daily")
            eurpln_row = quotes.latest_quote(conn, ids["eurpln"], granularity="daily")
            price_eur = price_row["close"] if price_row else None
            eurpln_rate = eurpln_row["close"] if eurpln_row else None

            history_rows = conn.execute(
                "SELECT date, market_value_eur, market_value_pln FROM portfolio_history "
                "ORDER BY date").fetchall()
            # Krok 28.1: kurs PLN/EUR danego dnia wyprowadzony z JUŻ POBRANEGO wiersza
            # `portfolio_history` (ten sam `rate`, którego użył `history.py::rebuild()`
            # do policzenia `market_value_pln`) — zero dodatkowych zapytań do NBP,
            # zero rozjazdu metodologii między krzywą portfela a krzywą benchmarku.
            rate_by_date = {
                r["date"]: r["market_value_pln"] / r["market_value_eur"]
                for r in history_rows
                if r["market_value_eur"] not in (None, 0) and r["market_value_pln"] is not None}

            xirr_pct = twr_pct = attribution = None
            benchmark_today_eur = benchmark_today_pln = None
            sharpe = volatility_pct = None
            xirr_flows: list[tuple[str, float]] = []

            # Krok 32: `daily_values`/`twr_flows` budowane bezwarunkowo (nie
            # potrzebują dzisiejszej ceny/kursu) — `max_drawdown()` działa
            # nawet, gdy dzisiejszy poll jeszcze się nie wykonał, tak jak
            # krzywa wartości i tabela rok-po-roku niżej w tej samej funkcji.
            # `twr_flows` (0.16.1): netuje kontrybucje/wypłaty (vesting, ESPP,
            # sprzedaże) z dziennych zwrotów — bez tego jeden dzień vestingu
            # wygląda jak +1000% "zwrotu" (bug znaleziony na produkcji 0.16.0).
            daily_values = [(r["date"], r["market_value_eur"]) for r in history_rows
                            if r["market_value_eur"] is not None]
            twr_flows = analytics_returns.build_twr_cashflows(conn)
            max_dd_result = analytics_risk.max_drawdown(daily_values, twr_flows)
            max_dd_pct = max_dd_result * 100 if max_dd_result is not None else None

            # Krzywa wartości i tabela rok-po-roku wyłącznie z `portfolio_history`
            # (przeliczana nocnym jobem) — NIEZALEŻNE od dzisiejszej ceny/kursu,
            # więc renderują się nawet gdy dzisiejszy poll jeszcze nie zdążył.
            chart_points = [
                {"date": r["date"], "value_pln": r["market_value_pln"],
                 "value_eur": r["market_value_eur"],
                 "benchmark_pln": None, "benchmark_eur": None}
                for r in history_rows]
            years: dict[str, dict] = {}
            for r in history_rows:
                y = r["date"][:4]
                years.setdefault(y, {"start_pln": r["market_value_pln"],
                                     "start_eur": r["market_value_eur"]})
                years[y]["end_pln"] = r["market_value_pln"]
                years[y]["end_eur"] = r["market_value_eur"]
            yearly_returns = []
            for y in sorted(years):
                start, end = years[y]["start_pln"], years[y]["end_pln"]
                pct = ((end - start) / start * 100) if start else None
                yearly_returns.append({
                    "year": y, "start_pln": start, "end_pln": end,
                    "start_eur": years[y]["start_eur"], "end_eur": years[y]["end_eur"],
                    "pct": pct})

            if price_eur and eurpln_rate:
                today = date.today().isoformat()
                current_qty = sum(r["qty_remaining"] for r in taxlots.open_lots(conn))
                xirr_flows = analytics_returns.build_xirr_cashflows(
                    conn, today, current_qty, price_eur)
                xirr_result = analytics_returns.xirr(xirr_flows)
                xirr_pct = xirr_result * 100 if xirr_result is not None else None

                twr_result = analytics_returns.twr(daily_values, twr_flows)
                twr_pct = twr_result * 100 if twr_result is not None else None

                sharpe = analytics_risk.sharpe_ratio(
                    daily_values, cfg["risk_free_rate_pct"], cashflows=twr_flows)
                volatility_result = analytics_risk.volatility_annualized(
                    daily_values, cashflows=twr_flows)
                volatility_pct = volatility_result * 100 if volatility_result is not None else None

                attribution = analytics_attribution.decompose(conn, price_eur, eurpln_rate)

                # Krok 28.1 — poprawka błędu: `xirr_flows` to gotówka w EUR
                # (`build_xirr_cashflows` używa `price_eur`/`net_received_eur`), a
                # OMXH25 jest zarejestrowany z `currency="EUR"` (main.py, ensure_instrument)
                # — `counterfactual()` zawsze liczył w EUR. Wcześniej wynik trafiał
                # do zmiennej `benchmark_today_pln` i renderował się z jednostką „zł"
                # bez żadnej konwersji: liczba w EUR pokazywana jako PLN (różnica
                # rzędu kursu EUR/PLN, ~4x zawyżenie/zaniżenie). Poprawka: liczyć
                # jawnie EUR, dokładać PLN przez ten sam kurs co reszta strony.
                benchmark_today_eur = analytics_benchmark.counterfactual(
                    conn, xirr_flows, ids["omxh25"])
                benchmark_today_pln = (
                    benchmark_today_eur * eurpln_rate if benchmark_today_eur is not None else None)

                bench_series_eur = dict(analytics_benchmark.counterfactual_series(
                    conn, xirr_flows, ids["omxh25"], [r["date"] for r in history_rows]))
                for point in chart_points:
                    b_eur = bench_series_eur.get(point["date"])
                    point["benchmark_eur"] = b_eur
                    rate = rate_by_date.get(point["date"])
                    point["benchmark_pln"] = b_eur * rate if b_eur is not None and rate is not None else None

            return render_template(
                "results.html", active="wyniki", version=__version__,
                xirr_pct=xirr_pct, twr_pct=twr_pct, attribution=attribution,
                benchmark_today_pln=benchmark_today_pln, benchmark_today_eur=benchmark_today_eur,
                sharpe=sharpe, volatility_pct=volatility_pct, max_dd_pct=max_dd_pct,
                chart_points=chart_points,
                yearly_returns=yearly_returns, has_history=bool(history_rows),
                print_mode=request.args.get("print") == "1")
        finally:
            conn.close()

    @app.get("/plan")
    def plan_get():
        """Krok 26 (docs/PLAN_KROK_26_doradca.md): cztery pytania, na które żadne
        narzędzie premium nie odpowiada — ile tracę sprzedając dziś, kiedy co wpada,
        ile da mi wpłacanie X EUR/mc, czy nie mam za dużo w jednym koszyku, który jest
        jednocześnie moim pracodawcą. Strona i sensor MQTT (`sensors.advisor_values`)
        liczą przez tę samą `advisor.overview()`, żeby nigdy nie pokazały dwóch różnych
        liczb dla tego samego faktu."""
        conn = _conn()
        try:
            ids = _ids(conn)
            price_row = quotes.latest_quote(conn, ids["primary"], granularity="daily")
            eurpln_row = quotes.latest_quote(conn, ids["eurpln"], granularity="daily")
            price_eur = price_row["close"] if price_row else None
            eurpln_rate = eurpln_row["close"] if eurpln_row else None

            cfg = settingsm.get_settings(conn)
            plan_overview = advisorm.overview(conn, cfg, price_eur, eurpln_rate)

            espp_result = None
            espp_error = None
            monthly_raw = request.args.get("espp_monthly")
            months_raw = request.args.get("espp_months")
            price_raw = request.args.get("espp_price")
            if monthly_raw and months_raw and price_raw:
                try:
                    espp_result = advisorm.espp_plan(
                        float(monthly_raw), int(float(months_raw)), float(price_raw),
                        eurpln_rate=eurpln_rate, match_pct=cfg["espp_match_pct"],
                        cost_basis_policy=cfg["cost_basis_policy"],
                        tax_pct=cfg["pl_capital_gains_tax_pct"])
                except ValueError as e:
                    espp_error = str(e)

            espp_scenarios = None
            if price_eur:
                espp_scenarios = [
                    ("bieżąca", price_eur), ("−20%", price_eur * 0.8),
                    ("+20%", price_eur * 1.2)]

            timing_result = None
            timing_qty_raw = request.args.get("timing_qty")
            timing_price_raw = request.args.get("timing_price")
            if timing_qty_raw and timing_price_raw:
                timing_result = advisorm.optimize_sale_timing(
                    conn, cfg, float(timing_qty_raw), float(timing_price_raw),
                    eurpln_rate=eurpln_rate)

            exit_result = None
            exit_error = None
            exit_qty_raw = request.args.get("exit_qty")
            exit_freq_raw = request.args.get("exit_freq")
            exit_periods_raw = request.args.get("exit_periods")
            if exit_qty_raw and exit_freq_raw and exit_periods_raw:
                try:
                    exit_result = advisorm.exit_plan(
                        conn, cfg, float(exit_qty_raw), exit_freq_raw,
                        int(float(exit_periods_raw)), price_eur or 0.0,
                        eurpln_rate=eurpln_rate)
                except (ValueError, taxlots.InsufficientLotsError) as e:
                    exit_error = str(e)

            return render_template(
                "plan.html", active="plan", version=__version__,
                overview=plan_overview, cfg=cfg, price_eur=price_eur,
                espp_result=espp_result, espp_error=espp_error,
                espp_monthly=monthly_raw, espp_months=months_raw, espp_price=price_raw,
                espp_scenarios=espp_scenarios,
                timing_result=timing_result, timing_qty=timing_qty_raw, timing_price=timing_price_raw,
                exit_result=exit_result, exit_error=exit_error,
                exit_qty=exit_qty_raw, exit_freq=exit_freq_raw, exit_periods=exit_periods_raw,
                has_restricted=bool(plan_overview["forfeit"]["items"]),
                has_timeline=bool(plan_overview["timeline"]["tranches"]),
                print_mode=request.args.get("print") == "1")
        finally:
            conn.close()

    @app.get("/api/preview/espp")
    def preview_espp():
        conn = _conn()
        try:
            try:
                monthly_eur = float(request.args.get("espp_monthly") or 0)
                months = int(float(request.args.get("espp_months") or 0))
                price_eur = float(request.args.get("espp_price") or 0)
            except ValueError:
                return {"ok": False, "error": "Niepoprawna liczba."}

            cfg = settingsm.get_settings(conn)
            ids = _ids(conn)
            eurpln_row = quotes.latest_quote(conn, ids["eurpln"], granularity="daily")
            eurpln_rate = eurpln_row["close"] if eurpln_row else None

            try:
                result = advisorm.espp_plan(
                    monthly_eur, months, price_eur, eurpln_rate=eurpln_rate,
                    match_pct=cfg["espp_match_pct"], cost_basis_policy=cfg["cost_basis_policy"],
                    tax_pct=cfg["pl_capital_gains_tax_pct"])
            except ValueError as e:
                return {"ok": False, "error": str(e)}

            lines = [
                {"label": "Akcje własne", "value": result["own_shares"], "unit": "szt."},
                {"label": "Akcje dopasowania", "value": result["matched_shares"], "unit": "szt."},
                {"label": "Razem", "value": result["total_shares"], "unit": "szt."},
            ]
            if result["tax_pln"] is not None:
                lines.append({"label": "Podatek", "value": result["tax_pln"], "unit": "PLN"})
                lines.append({"label": "Na rękę", "value": result["net_proceeds_pln"],
                              "unit": "PLN", "emphasis": True})
            return {"ok": True, "lines": lines}
        finally:
            conn.close()

    @app.get("/api/preview/sale-timing")
    def preview_sale_timing():
        conn = _conn()
        try:
            try:
                quantity = float(request.args.get("timing_qty") or 0)
                price_eur = float(request.args.get("timing_price") or 0)
            except ValueError:
                return {"ok": False, "error": "Niepoprawna liczba."}
            if quantity <= 0 or price_eur <= 0:
                return {"ok": False, "error": "Ilość i cena muszą być dodatnie."}

            cfg = settingsm.get_settings(conn)
            ids = _ids(conn)
            eurpln_row = quotes.latest_quote(conn, ids["eurpln"], granularity="daily")
            eurpln_rate = eurpln_row["close"] if eurpln_row else None

            result = advisorm.optimize_sale_timing(
                conn, cfg, quantity, price_eur, eurpln_rate=eurpln_rate)

            if result["today"] is None or result["jan2_next_year"] is None:
                return {"ok": False, "error": "Brak pokrycia lotami dla jednego ze scenariuszy."}

            lines = [
                {"label": "Podatek dziś (po stracie)",
                 "value": result["today"]["tax_with_max_loss_pln"], "unit": "PLN"},
                {"label": "Podatek 2 stycznia (po stracie)",
                 "value": result["jan2_next_year"]["tax_with_max_loss_pln"], "unit": "PLN"},
                {"label": "Różnica netto (podatek + przepadek)",
                 "value": result["delta_total_pln"], "unit": "PLN", "emphasis": True},
            ]
            return {"ok": True, "lines": lines}
        finally:
            conn.close()

    @app.get("/api/preview/exit-plan")
    def preview_exit_plan():
        conn = _conn()
        try:
            try:
                shares_per_period = float(request.args.get("exit_qty") or 0)
                frequency = request.args.get("exit_freq") or ""
                num_periods = int(float(request.args.get("exit_periods") or 0))
            except ValueError:
                return {"ok": False, "error": "Niepoprawna liczba."}

            cfg = settingsm.get_settings(conn)
            ids = _ids(conn)
            price_row = quotes.latest_quote(conn, ids["primary"], granularity="daily")
            eurpln_row = quotes.latest_quote(conn, ids["eurpln"], granularity="daily")
            price_eur = price_row["close"] if price_row else None
            eurpln_rate = eurpln_row["close"] if eurpln_row else None

            try:
                result = advisorm.exit_plan(
                    conn, cfg, shares_per_period, frequency, num_periods,
                    price_eur or 0.0, eurpln_rate=eurpln_rate)
            except (ValueError, taxlots.InsufficientLotsError) as e:
                return {"ok": False, "error": str(e)}

            lines = [
                {"label": "Łącznie sprzedanych akcji",
                 "value": result["totals"]["shares_sold"], "unit": "szt."},
            ]
            if result["totals"]["tax_pln"] is not None:
                lines.append({"label": "Podatek łącznie",
                              "value": result["totals"]["tax_pln"], "unit": "PLN"})
                lines.append({"label": "Na rękę", "value": result["totals"]["net_proceeds_pln"],
                              "unit": "PLN", "emphasis": True})
            return {"ok": True, "lines": lines}
        finally:
            conn.close()

    @app.get("/api/preview/copilot")
    def preview_copilot():
        """Krok 33 (docs/PLAN_KROK_33_copilot.md): podgląd co-pilota BEZ
        skutków ubocznych — nie woła AI, nie wysyła powiadomienia i nie
        zapisuje do alerts_log (nie konsumuje cooldownu), więc wolno go
        wołać na produkcji do weryfikacji. `?today=YYYY-MM-DD` pozwala
        sprawdzić, co odpaliłoby się innego dnia."""
        conn = _conn()
        try:
            today = request.args.get("today") or None
            cfg = settingsm.get_settings(conn)
            try:
                return ai_copilot.preview(conn, cfg, today=today)
            except ValueError:
                return {"ok": False, "error": "Niepoprawna data (oczekiwano RRRR-MM-DD)."}
        finally:
            conn.close()

    def _assistant_ask_and_redirect(question: str):
        """Wspólna ścieżka dla POST /asystent i GET /asystent?q= (pole
        szybkiego pytania na pulpicie, krok 29.7) — OBIE kończą się
        przekierowaniem do CZYSTEGO /asystent, żeby odświeżenie strony po
        odpowiedzi nie powtórzyło pytania (to samo uzasadnienie co
        POST-redirect-GET gdzie indziej w aplikacji). Odpowiedź trafia do
        chat_log i pojawia się jako pierwszy wiersz historii na kolejnym GET."""
        conn = _conn()
        try:
            cfg = dict(settingsm.get_settings(conn), **_ai_keys())
            if question and cfg.get("ai_chat_enabled", 1):
                ai_chat.ask(conn, cfg, question)
            return redirect(url_for("assistant_get"))
        finally:
            conn.close()

    @app.get("/asystent")
    def assistant_get():
        q = request.args.get("q", "").strip()
        if q:
            return _assistant_ask_and_redirect(q)
        conn = _conn()
        try:
            cfg = dict(settingsm.get_settings(conn), **_ai_keys())
            return render_template(
                "assistant.html", active="assistant", version=__version__,
                chat_enabled=bool(cfg.get("ai_chat_enabled", 1)),
                history=ai_chat.history(conn, limit=20),
                ai_status=ai_status.snapshot(conn, cfg))
        finally:
            conn.close()

    @app.post("/asystent")
    def assistant_post():
        return _assistant_ask_and_redirect(request.form.get("question", "").strip())

    @app.route("/api/asystent", methods=["GET", "POST"])
    def assistant_api():
        """JSON dla ewentualnej przyszłej wersji z JS — CELOWO bez auto-fire
        na wpisywanie (initFormPreview() debounce'uje na 'input', co dla
        pytań w naturalnym języku wołałoby AI na każde naciśnięcie klawisza;
        ten endpoint wymaga jawnego wywołania, nie jest dziś podpięty pod
        żaden formularz z debounce)."""
        conn = _conn()
        try:
            cfg = dict(settingsm.get_settings(conn), **_ai_keys())
            if request.method == "POST":
                body = request.get_json(silent=True) or {}
                question = (body.get("question") or request.form.get("question", "")).strip()
            else:
                question = request.args.get("q", "").strip()
            if not cfg.get("ai_chat_enabled", 1):
                return {"ok": False, "intent": None, "params": {}, "title": None,
                       "lines": [], "detail_url": None,
                       "error": "Asystent jest wyłączony w Ustawieniach.", "answer_pl": None}
            return ai_chat.ask(conn, cfg, question)
        finally:
            conn.close()

    @app.get("/imports")
    def imports_get():
        conn = _conn()
        try:
            history = conn.execute(
                "SELECT * FROM imports ORDER BY imported_at DESC").fetchall()
            conflict_rows = conn.execute(
                "SELECT * FROM import_conflicts WHERE resolved = 0 ORDER BY id DESC"
            ).fetchall()
            conflicts = []
            for r in conflict_rows:
                d = dict(r)
                d["existing"] = json.loads(d["existing_json"]) if d["existing_json"] else {}
                d["incoming"] = json.loads(d["incoming_json"]) if d["incoming_json"] else {}
                conflicts.append(d)
            return render_template(
                "imports.html", active="imports", version=__version__,
                history=[dict(r) for r in history], conflicts=conflicts,
                report=request.args.get("report"), sold=request.args.get("sold") == "1",
                error=request.args.get("error"))
        finally:
            conn.close()

    @app.post("/imports/upload")
    def imports_upload():
        conn = _conn()
        try:
            uploaded = request.files.get("pdf_file")
            if not uploaded or not uploaded.filename:
                return redirect(url_for("imports_get"))
            data = uploaded.read()
            cfg = settingsm.get_settings(conn)
            with dbm.WRITE_LOCK:
                report = computershare_pdf.import_statement(
                    conn, data, uploaded.filename, cfg)
                grantsm.reconcile_vesting(conn)
            return redirect(url_for(
                "imports_get",
                report=f"{report['rows_inserted']}/{report['rows_unchanged']}/"
                       f"{report['rows_conflict']}"))
        finally:
            conn.close()

    @app.post("/imports/conflicts/<int:conflict_id>/resolve")
    def imports_resolve_conflict(conflict_id: int):
        conn = _conn()
        try:
            resolution = request.form.get("resolution", "")
            with dbm.WRITE_LOCK:
                conn.execute(
                    "UPDATE import_conflicts SET resolved = 1, resolution = ? WHERE id = ?",
                    (resolution, conflict_id))
                conn.commit()
            return redirect(url_for("imports_get"))
        finally:
            conn.close()

    @app.post("/imports/conflicts/<int:conflict_id>/confirm-sale")
    def imports_confirm_sale(conflict_id: int):
        """Zatwierdza jednym kliknięciem sprzedaż wykrytą w PDF (Withhold-to-Cover Typ B /
        „Sell (Shares)"), bez ręcznego przepisywania liczb do /lots/sell — czyta dane wprost
        z `incoming_json` zapisanego przy imporcie (krok 13)."""
        conn = _conn()
        try:
            row = conn.execute(
                "SELECT * FROM import_conflicts WHERE id = ?", (conflict_id,)).fetchone()
            if row is None:
                return redirect(url_for("imports_get"))
            incoming = json.loads(row["incoming_json"])
            try:
                with dbm.WRITE_LOCK:
                    sale_id = taxlots.record_sale(
                        conn, incoming["execution_date"], incoming["quantity"],
                        incoming["sale_price_eur"], fee_eur=incoming.get("fees_eur", 0.0),
                        proceeds_eur=incoming.get("sale_proceeds_eur"))
                    conn.execute(
                        "UPDATE import_conflicts SET resolved = 1, resolution = ? WHERE id = ?",
                        (f"zaksięgowano automatycznie jako sprzedaż (sale_id={sale_id})",
                         conflict_id))
                    conn.commit()
                return redirect(url_for("imports_get", sold="1"))
            except (taxlots.InsufficientLotsError, taxlots.CostBasisMissingError) as e:
                return redirect(url_for("imports_get", error=str(e)))
        finally:
            conn.close()

    def _restore_dir() -> Path:
        d = Path(db_path).parent / "tmp_restore"
        d.mkdir(parents=True, exist_ok=True)
        return d

    def _cleanup_stale_restore_files(restore_dir: Path, keep_token: str | None) -> None:
        """Pliki podglądu przywracania zawierają pełną kopię bazy (dane
        podatkowe) — nie zostają na dysku bezterminowo, jeśli użytkownik
        wgra plik i nigdy nie potwierdzi ani nie odrzuci podglądu."""
        cutoff = datetime.now().timestamp() - 3600
        for f in restore_dir.glob("*.zip"):
            if keep_token and f.stem == keep_token:
                continue
            if f.stat().st_mtime < cutoff:
                f.unlink(missing_ok=True)

    def _backup_dir() -> Path:
        return Path(os.environ.get("BACKUP_SHARE", "/share/nokia_tracker")) / "backup"

    @app.get("/dane")
    def data_get():
        conn = _conn()
        try:
            token = request.args.get("token")
            restore_dir = _restore_dir()
            _cleanup_stale_restore_files(restore_dir, token)

            preview = None
            preview_error = None
            if token:
                zip_path = restore_dir / f"{token}.zip"
                if zip_path.exists():
                    try:
                        preview = backupm.restore_preview(db_path, zip_path.read_bytes())
                    except backupm.IncompatibleBackupError as e:
                        preview_error = str(e)
                else:
                    preview_error = "Podgląd wygasł — wgraj plik ponownie."

            conflicts_count = conn.execute(
                "SELECT COUNT(*) FROM import_conflicts WHERE resolved = 0").fetchone()[0]

            backup_dir = _backup_dir()
            last_snapshot = None
            if backup_dir.is_dir():
                snapshots = sorted(backup_dir.glob("nokia_*.zip"))
                if snapshots:
                    stat = snapshots[-1].stat()
                    last_snapshot = {
                        "name": snapshots[-1].name,
                        "size_kb": round(stat.st_size / 1024, 1),
                        "mtime": datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M"),
                    }

            return render_template(
                "data.html", active="data", version=__version__,
                token=token, preview=preview, preview_error=preview_error,
                conflicts_count=conflicts_count, last_snapshot=last_snapshot,
                restored=request.args.get("restored") == "1",
                error=request.args.get("error"))
        finally:
            conn.close()

    @app.get("/dane/eksport.zip")
    def data_export_zip():
        data = backupm.export_zip(db_path)
        filename = f"nokia_tracker_{date.today().isoformat()}.zip"
        return Response(
            data, mimetype="application/zip",
            headers={"Content-Disposition": f"attachment; filename={filename}"})

    @app.post("/dane/import/preview")
    def data_import_preview():
        uploaded = request.files.get("backup_file")
        if not uploaded or not uploaded.filename:
            return redirect(url_for("data_get"))
        token = secrets.token_urlsafe(16)
        (_restore_dir() / f"{token}.zip").write_bytes(uploaded.read())
        return redirect(url_for("data_get", token=token))

    @app.post("/dane/import/confirm")
    def data_import_confirm():
        token = request.form.get("token", "")
        zip_path = _restore_dir() / f"{token}.zip"
        if not token or not zip_path.exists():
            return redirect(url_for("data_get", error="Podgląd wygasł — wgraj plik ponownie."))
        try:
            with dbm.WRITE_LOCK:
                backupm.restore_apply(db_path, zip_path.read_bytes())
        except backupm.IncompatibleBackupError as e:
            return redirect(url_for("data_get", error=str(e)))
        finally:
            zip_path.unlink(missing_ok=True)
        return redirect(url_for("data_get", restored="1"))

    @app.get("/news")
    def news_page():
        conn = _conn()
        try:
            # Krok 18: `s.horizon`/`s.tags` odpytywane tu wcześniej, ale szablon
            # nigdy ich nie renderował — martwa robota na każde żądanie, usunięte.
            # `ns.kind` (rss/gdelt/finnhub/marketaux) dołożone jako kolumna „Źródło" —
            # dotąd niewidoczne w UI mimo że news.py śledzi providera per wpis.
            rows = conn.execute(
                "SELECT n.*, s.sentiment, s.impact, s.thesis_pl, ns.kind AS source_kind "
                "FROM news n LEFT JOIN news_scores s ON s.news_id = n.id "
                "LEFT JOIN news_sources ns ON ns.id = n.source_id "
                # Krok 28.6: 50 -> 200 — limit istniał wyłącznie żeby nie
                # renderować bez końca; teraz "Pokaż więcej" (news.html/app.js)
                # ujawnia po 20 client-side, więc wyższy limit ma sens dopiero
                # z paginacją po stronie serwera. 200 to wciąż jedno zapytanie
                # SQLite po indeksowanej kolumnie, bez odczuwalnego kosztu.
                "ORDER BY n.published_at DESC LIMIT 200"
            ).fetchall()
            items = [dict(r) for r in rows]
            return render_template("news.html", active="news", version=__version__, items=items)
        finally:
            conn.close()

    @app.get("/forecasts")
    def forecasts_page():
        conn = _conn()
        try:
            rows = conn.execute(
                "SELECT * FROM forecasts ORDER BY created_at DESC LIMIT 60"
            ).fetchall()
            # Krok 18: jedyna liczba, po którą się tu przychodzi ("czy prognozy się
            # sprawdzają") żyła dotąd tylko na pulpicie (`sensors.forecast_values`);
            # ta strona pokazywała samą historię bez podsumowania.
            accuracy_pct = sensors.forecast_values(conn).get("forecast_accuracy_pct")
            return render_template(
                "forecasts.html", active="forecasts", version=__version__,
                items=[dict(r) for r in rows], accuracy_pct=accuracy_pct)
        finally:
            conn.close()

    @app.get("/settings")
    def settings_get():
        conn = _conn()
        try:
            cfg = settingsm.get_settings(conn)
            keys = _ai_keys()
            local_models = (openai_compat.list_models(cfg["local_llm_base_url"],
                                                       keys["local_llm_api_key"])
                            if cfg["local_llm_base_url"] else [])
            return render_template(
                "settings.html", active="settings", version=__version__, cfg=cfg,
                local_models=local_models, saved=request.args.get("saved") == "1",
                ai_status=ai_status.snapshot(conn, dict(cfg, **_ai_keys()), local_models))
        finally:
            conn.close()

    @app.post("/settings")
    def settings_post():
        conn = _conn()
        try:
            updates = {
                "ai_primary": request.form.get("ai_primary", "local"),
                "ai_fallback": request.form.get("ai_fallback", "gemini"),
                "local_llm_model": request.form.get("local_llm_model", ""),
                "gemini_model": request.form.get("gemini_model", ""),
                "anthropic_model": request.form.get("anthropic_model", ""),
                "ai_recommendations_enabled": 1 if request.form.get("ai_recommendations_enabled") else 0,
                "ai_chat_enabled": 1 if request.form.get("ai_chat_enabled") else 0,
                "ai_chat_narration_enabled": 1 if request.form.get("ai_chat_narration_enabled") else 0,
                "ai_max_calls_per_day_local": int(
                    request.form.get("ai_max_calls_per_day_local") or 500),
                "alert_sentiment_drop": float(request.form.get("alert_sentiment_drop") or 0.5),
                "alert_price_move_pct": float(request.form.get("alert_price_move_pct") or 3.0),
                "alert_on_forecast_break": 1 if request.form.get("alert_on_forecast_break") else 0,
                "alert_min_interval_minutes": int(
                    request.form.get("alert_min_interval_minutes") or 120),
                "notify_service": request.form.get("notify_service", ""),
                "notify_news_enabled": 1 if request.form.get("notify_news_enabled") else 0,
                "notify_news_min_impact": int(
                    request.form.get("notify_news_min_impact") or 1),
                "notify_digest_enabled": 1 if request.form.get("notify_digest_enabled") else 0,
                "digest_time": request.form.get("digest_time", "20:10"),
                "cost_basis_policy": request.form.get("cost_basis_policy", "own_only"),
                # Krok 18: dawniej tylko odczytywalne (domyślne 35% pokazywane jako
                # tekst na /dywidendy bez możliwości zmiany) — te cztery stawki
                # wpływają na każdą kwotę PLN w aplikacji, więc trafiają do UI.
                "finnish_withholding_pct": float(
                    request.form.get("finnish_withholding_pct") or 35.0),
                "treaty_withholding_pct": float(
                    request.form.get("treaty_withholding_pct") or 15.0),
                "pl_capital_gains_tax_pct": float(
                    request.form.get("pl_capital_gains_tax_pct") or 19.0),
                "tax_year": int(request.form.get("tax_year") or 0),
                # Krok 26: doradca planu pracowniczego.
                "other_net_worth_pln": float(
                    request.form.get("other_net_worth_pln") or 0.0),
                "concentration_alert_pct": float(
                    request.form.get("concentration_alert_pct") or 25.0),
            }
            with dbm.WRITE_LOCK:
                settingsm.set_settings(conn, updates)
            return redirect(url_for("settings_get", saved="1"))
        finally:
            conn.close()

    def _pit38_report_for_request(conn):
        """Wspólne dla /pit38 i obu eksportów: rok z ?year= (domyślnie
        cfg['tax_year'] lub bieżący rok) + roczny raport."""
        cfg = settingsm.get_settings(conn)
        taxdiv.backfill_pl_tax_due(conn, cfg)
        year = request.args.get("year", type=int) or cfg.get("tax_year") or datetime.now().year
        report = taxpit38.annual_report(conn, cfg, year)
        return cfg, year, report

    def _enrich_trace_row_for_export(conn, row: dict) -> dict:
        """Krok 16 (§8.4): dokłada do wiersza `report['sale_trace']` kwoty EUR
        (pochodne zamrożonego PLN — `pln / kurs`, więc zawsze spójne z tym,
        co widać na ekranie) i numery tabel NBP — eksport ma być tym samym
        dowodem co `/pit38`, nie jego uboższą wersją. Zero nowych zapytań do
        NBP: `table_no_for_effective_date` czyta tylko lokalną `nbp_rates`."""
        lot_rate = row["lot_nbp_rate"]
        sale_rate = row["sale_nbp_rate"]
        return {
            **row,
            "cost_eur": row["cost_pln"] / lot_rate if lot_rate else None,
            "revenue_eur": row["revenue_pln"] / sale_rate if sale_rate else None,
            "lot_table_no": (fx_nbp.table_no_for_effective_date(conn, row["lot_nbp_rate_date"])
                             if row["lot_nbp_rate_date"] else None),
            "sale_table_no": (fx_nbp.table_no_for_effective_date(conn, row["sale_nbp_rate_date"])
                              if row["sale_nbp_rate_date"] else None),
        }

    @app.context_processor
    def _inject_nav_tax_year():
        """Krok 28.3 (docs/PLAN_KROK_28_ux_mobile.md §3): jeden selektor roku w
        nagłówku (base.html) zamiast trzech osobnych kopii (`pit38.html`,
        `sales.html`, `wizard.html`). Submituje `?year=` na BIEŻĄCĄ stronę
        (`request.path`) — trasy, które już czytają `request.args.get("year")`
        (pit38/sales/wizard), dostają go bez zmiany własnej logiki; trasy, które
        nie znają tego parametru, po prostu go ignorują (Flask nie waliduje
        nieznanych query stringów) — zero ryzyka regresji na stronach, które
        dziś selektora nie mają. Pusty wybór ("wszystkie") wysyła `year=` — dla
        `request.args.get("year", type=int)` to `None`, czyli dokładnie ten sam
        fallback co brak parametru w ogóle (sales: wszystkie lata; pit38/wizard:
        `cfg['tax_year']` albo rok bieżący)."""
        conn = _conn()
        try:
            return {"nav_tax_years": taxpit38.years_with_data(conn),
                    "nav_selected_year": request.args.get("year", type=int)}
        finally:
            conn.close()

    @app.get("/pit38")
    def pit38_get():
        conn = _conn()
        try:
            cfg, year, report = _pit38_report_for_request(conn)

            whatif_result = None
            whatif_error = None
            qty_raw = request.args.get("whatif_qty")
            price_raw = request.args.get("whatif_price")
            if qty_raw and price_raw:
                try:
                    whatif_result = taxwhatif.simulate_sale(
                        conn, cfg, float(qty_raw), float(price_raw))
                except (taxlots.InsufficientLotsError, taxlots.CostBasisMissingError) as e:
                    whatif_error = str(e)

            current_price = sensors.market_values(conn, _ids(conn)["primary"]).get("price_eur")

            # Krok 28.4 (docs/PLAN_KROK_28_ux_mobile.md §4): waterfall PIT-38 —
            # WYŁĄCZNIE Poz. C (przychód ze sprzedaży), zero nowej matematyki,
            # tylko reshaping już policzonych pól `report`. "Strata odliczona" to
            # segment INFORMACYJNY (pokazuje wielkość tarczy), nie wchodzi do
            # łańcucha przychód-koszt-podatek-na rękę — strata obniża PODATEK, nie
            # jest realnym wypływem gotówki, więc doliczanie jej do łańcucha
            # dawałoby "na rękę" niezgodne z prawdziwą kwotą (dochód - podatek).
            active_data = report["policies"][cfg["cost_basis_policy"]]
            loss_cf = report.get("loss_carryforward") or {}
            has_loss = bool(loss_cf.get("items"))
            tax_after_loss = loss_cf["tax_after_loss_pln"] if has_loss else active_data["tax_pln"]
            loss_used = loss_cf["total_used_this_year_pln"] if has_loss else 0.0
            net_in_hand = active_data["income_pln"] - tax_after_loss
            waterfall_pit38 = {
                "revenue": active_data["revenue_pln"], "cost": active_data["cost_pln"],
                "income": active_data["income_pln"], "loss_used": loss_used,
                "tax": tax_after_loss, "net": net_in_hand,
            }

            return render_template(
                "pit38.html", active="pit38", version=__version__,
                year=year, report=report, cfg=cfg,
                whatif_result=whatif_result, whatif_error=whatif_error,
                whatif_qty=qty_raw, whatif_price=price_raw,
                current_price=current_price, waterfall_pit38=waterfall_pit38,
                print_mode=request.args.get("print") == "1")
        finally:
            conn.close()

    @app.get("/pit38/export.csv")
    def pit38_export_csv():
        conn = _conn()
        try:
            _cfg, year, report = _pit38_report_for_request(conn)

            buf = io.StringIO()
            w = csv.writer(buf)
            w.writerow(["PIT-38", year])
            w.writerow([])
            w.writerow(["Polityka", "Przychod PLN", "Koszt PLN", "Dochod PLN", "Podatek PLN"])
            for name, data in report["policies"].items():
                w.writerow([name, data["revenue_pln"], data["cost_pln"],
                            data["income_pln"], data["tax_pln"]])
            w.writerow([])
            w.writerow(["Sekcja G (dywidendy)"])
            w.writerow(["Liczba dywidend", report["section_g"]["dividend_count"]])
            w.writerow(["Brutto PLN", report["section_g"]["gross_pln"]])
            w.writerow(["Pobrane u zrodla PLN", report["section_g"]["withholding_paid_pln"]])
            w.writerow(["Zaliczenie traktatowe PLN", report["section_g"]["credit_pln"]])
            w.writerow(["Belka PLN", report["section_g"]["belka_pln"]])
            w.writerow(["Doplata w PL PLN", report["section_g"]["pl_tax_due_pln"]])
            w.writerow(["Do odzyskania z Vero PLN",
                        report["section_g"]["reclaimable_from_finland_pln"]])
            w.writerow([])
            w.writerow(["PIT/ZG", "Kraj", report["pit_zg"]["country"]])
            w.writerow(["Dochod zagraniczny PLN", report["pit_zg"]["foreign_income_pln"]])
            w.writerow(["Podatek zaplacony za granica PLN",
                        report["pit_zg"]["foreign_tax_paid_pln"]])
            w.writerow([])
            w.writerow(["Slad per lot"])
            w.writerow(["Lot ID", "Data nabycia", "Typ", "Ilosc", "Koszt EUR", "Koszt PLN",
                        "Przychod EUR", "Przychod PLN", "Kurs NBP lotu", "Data kursu lotu",
                        "Tabela NBP lotu", "Data sprzedazy", "Kurs NBP sprzedazy",
                        "Tabela NBP sprzedazy"])
            for raw_row in report["sale_trace"]:
                row = _enrich_trace_row_for_export(conn, raw_row)
                w.writerow([
                    row["lot_id"], row["acquired_date"], row["lot_type"], row["quantity"],
                    row["cost_eur"], row["cost_pln"], row["revenue_eur"], row["revenue_pln"],
                    row["lot_nbp_rate"], row["lot_nbp_rate_date"], row["lot_table_no"],
                    row["sale_date"], row["sale_nbp_rate"], row["sale_table_no"]])

            filename = f"pit38_{year}.csv"
            return Response(
                "﻿" + buf.getvalue(),
                mimetype="text/csv; charset=utf-8",
                headers={"Content-Disposition": f"attachment; filename={filename}"})
        finally:
            conn.close()

    @app.get("/pit38/export.xlsx")
    def pit38_export_xlsx():
        conn = _conn()
        try:
            _cfg, year, report = _pit38_report_for_request(conn)

            wb = openpyxl.Workbook()
            ws_summary = wb.active
            ws_summary.title = "Podsumowanie"
            ws_summary.append(
                ["Polityka", "Przychod PLN", "Koszt PLN", "Dochod PLN", "Podatek PLN"])
            for name, data in report["policies"].items():
                ws_summary.append([name, data["revenue_pln"], data["cost_pln"],
                                    data["income_pln"], data["tax_pln"]])
            ws_summary.append([])
            ws_summary.append(["Sekcja G (dywidendy)"])
            ws_summary.append(["Liczba dywidend", report["section_g"]["dividend_count"]])
            ws_summary.append(["Brutto PLN", report["section_g"]["gross_pln"]])
            ws_summary.append(["Doplata w PL PLN", report["section_g"]["pl_tax_due_pln"]])
            ws_summary.append(
                ["Do odzyskania z Vero PLN", report["section_g"]["reclaimable_from_finland_pln"]])
            ws_summary.append([])
            ws_summary.append(["PIT/ZG", "Kraj", report["pit_zg"]["country"]])
            ws_summary.append(["Dochod zagraniczny PLN", report["pit_zg"]["foreign_income_pln"]])

            ws_trace = wb.create_sheet("Ślad per lot")
            ws_trace.append(["Lot ID", "Data nabycia", "Typ", "Ilosc", "Koszt EUR",
                              "Koszt PLN", "Przychod EUR", "Przychod PLN", "Kurs NBP lotu",
                              "Data kursu lotu", "Tabela NBP lotu", "Data sprzedazy",
                              "Kurs NBP sprzedazy", "Tabela NBP sprzedazy"])
            for raw_row in report["sale_trace"]:
                row = _enrich_trace_row_for_export(conn, raw_row)
                ws_trace.append([
                    row["lot_id"], row["acquired_date"], row["lot_type"], row["quantity"],
                    row["cost_eur"], row["cost_pln"], row["revenue_eur"], row["revenue_pln"],
                    row["lot_nbp_rate"], row["lot_nbp_rate_date"], row["lot_table_no"],
                    row["sale_date"], row["sale_nbp_rate"], row["sale_table_no"]])

            ws_div = wb.create_sheet("Dywidendy")
            ws_div.append(["Rok", year])
            ws_div.append(["Liczba dywidend", report["section_g"]["dividend_count"]])
            ws_div.append(["Brutto PLN", report["section_g"]["gross_pln"]])
            ws_div.append(["Pobrane u zrodla PLN", report["section_g"]["withholding_paid_pln"]])
            ws_div.append(["Doplata w PL PLN", report["section_g"]["pl_tax_due_pln"]])
            ws_div.append(
                ["Do odzyskania z Vero PLN", report["section_g"]["reclaimable_from_finland_pln"]])

            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            filename = f"pit38_{year}.xlsx"
            return Response(
                buf.getvalue(),
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename={filename}"})
        finally:
            conn.close()

    @app.get("/pit38/kreator")
    def pit38_wizard_get():
        conn = _conn()
        try:
            cfg = settingsm.get_settings(conn)
            year = request.args.get("year", type=int) or cfg.get("tax_year") or datetime.now().year
            taxlosses.rebuild(conn, cfg)
            report = taxpit38.annual_report(conn, cfg, year)
            steps = taxlosses.wizard_steps(conn, cfg, year, report)
            closed = taxlosses.is_year_closed(conn, year)
            can_close = all(s["done"] for s in steps if s["key"] in ("import", "conflicts", "balance"))
            return render_template(
                "wizard.html", active="pit38", version=__version__,
                year=year, steps=steps, report=report,
                closed=closed, can_close=can_close, cfg=cfg,
                deduct_error=request.args.get("deduct_error"),
                close_error=request.args.get("close_error") == "1")
        finally:
            conn.close()

    @app.post("/pit38/kreator/odlicz")
    def pit38_wizard_deduct():
        conn = _conn()
        try:
            cfg = settingsm.get_settings(conn)
            year = int(request.form["year"])
            try:
                with dbm.WRITE_LOCK:
                    taxlosses.record_deduction(
                        conn, cfg, int(request.form["loss_id"]), year,
                        float(request.form["amount_pln"]))
            except ValueError as e:
                return redirect(url_for("pit38_wizard_get", year=year, deduct_error=str(e)))
            return redirect(url_for("pit38_wizard_get", year=year))
        finally:
            conn.close()

    @app.post("/pit38/kreator/zamknij")
    def pit38_wizard_close():
        conn = _conn()
        try:
            year = int(request.form["year"])
            cfg = settingsm.get_settings(conn)
            report = taxpit38.annual_report(conn, cfg, year)
            steps = taxlosses.wizard_steps(conn, cfg, year, report)
            blocking = [s for s in steps if s["key"] in ("import", "conflicts", "balance")
                        and not s["done"]]
            if blocking:
                return redirect(url_for("pit38_wizard_get", year=year, close_error="1"))
            with dbm.WRITE_LOCK:
                taxlosses.close_year(conn, cfg, year, report["total_due_pln"])
            return redirect(url_for("pit38_wizard_get", year=year, closed="1"))
        finally:
            conn.close()

    @app.post("/pit38/kreator/odblokuj")
    def pit38_wizard_reopen():
        conn = _conn()
        try:
            year = int(request.form["year"])
            with dbm.WRITE_LOCK:
                taxlosses.reopen_year(conn, year)
            return redirect(url_for("pit38_wizard_get", year=year))
        finally:
            conn.close()

    return app
