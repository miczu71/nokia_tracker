"""Serializacja PIT-38 do CSV/XLSX. Wyodrębnione z
`web.py::pit38_export_csv`/`pit38_export_xlsx` (E3 — docs/ROADMAP_V3.md) —
czysta serializacja już policzonego `report`, nie „widok" w sensie danych
domenowych, stąd `exports/`, nie `views/`. `to_csv` zwraca tekst BEZ BOM-u —
BOM jest sprawą dostawy HTTP (`Response`), doklejaną przez trasę razem z
mimetype."""
from __future__ import annotations

import csv
import io

import openpyxl


def to_csv(year: int, report: dict, trace_rows: list[dict]) -> str:
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["PIT-38", year])
    w.writerow([])
    w.writerow(["Polityka", "Przychod PLN", "Koszt PLN", "Dochod PLN", "Podatek PLN"])
    for name, data in report["policies"].items():
        w.writerow([name, data["revenue_pln"], data["cost_pln"],
                    data["income_pln"], data["tax_pln"]])
    w.writerow([])
    w.writerow(["Sekcja G (dywidendy)"])
    w.writerow(["Liczba dywidend", report["section_g"]["dividend_count"]])
    w.writerow(["Brutto PLN", report["section_g"]["gross_pln"]])
    w.writerow(["Pobrane u zrodla PLN", report["section_g"]["withholding_paid_pln"]])
    w.writerow(["Zaliczenie traktatowe PLN", report["section_g"]["credit_pln"]])
    w.writerow(["Belka PLN", report["section_g"]["belka_pln"]])
    w.writerow(["Doplata w PL PLN", report["section_g"]["pl_tax_due_pln"]])
    w.writerow(["Do odzyskania z Vero PLN",
                report["section_g"]["reclaimable_from_finland_pln"]])
    w.writerow([])
    w.writerow(["PIT/ZG", "Kraj", report["pit_zg"]["country"]])
    w.writerow(["Dochod zagraniczny PLN", report["pit_zg"]["foreign_income_pln"]])
    w.writerow(["Podatek zaplacony za granica PLN",
                report["pit_zg"]["foreign_tax_paid_pln"]])
    w.writerow([])
    w.writerow(["Slad per lot"])
    w.writerow(["Lot ID", "Data nabycia", "Typ", "Ilosc", "Koszt EUR", "Koszt PLN",
                "Przychod EUR", "Przychod PLN", "Kurs NBP lotu", "Data kursu lotu",
                "Tabela NBP lotu", "Data sprzedazy", "Kurs NBP sprzedazy",
                "Tabela NBP sprzedazy"])
    for row in trace_rows:
        w.writerow([
            row["lot_id"], row["acquired_date"], row["lot_type"], row["quantity"],
            row["cost_eur"], row["cost_pln"], row["revenue_eur"], row["revenue_pln"],
            row["lot_nbp_rate"], row["lot_nbp_rate_date"], row["lot_table_no"],
            row["sale_date"], row["sale_nbp_rate"], row["sale_table_no"]])
    return buf.getvalue()


def to_xlsx(year: int, report: dict, trace_rows: list[dict]) -> bytes:
    wb = openpyxl.Workbook()
    ws_summary = wb.active
    ws_summary.title = "Podsumowanie"
    ws_summary.append(
        ["Polityka", "Przychod PLN", "Koszt PLN", "Dochod PLN", "Podatek PLN"])
    for name, data in report["policies"].items():
        ws_summary.append([name, data["revenue_pln"], data["cost_pln"],
                            data["income_pln"], data["tax_pln"]])
    ws_summary.append([])
    ws_summary.append(["Sekcja G (dywidendy)"])
    ws_summary.append(["Liczba dywidend", report["section_g"]["dividend_count"]])
    ws_summary.append(["Brutto PLN", report["section_g"]["gross_pln"]])
    ws_summary.append(["Doplata w PL PLN", report["section_g"]["pl_tax_due_pln"]])
    ws_summary.append(
        ["Do odzyskania z Vero PLN", report["section_g"]["reclaimable_from_finland_pln"]])
    ws_summary.append([])
    ws_summary.append(["PIT/ZG", "Kraj", report["pit_zg"]["country"]])
    ws_summary.append(["Dochod zagraniczny PLN", report["pit_zg"]["foreign_income_pln"]])

    ws_trace = wb.create_sheet("Ślad per lot")
    ws_trace.append(["Lot ID", "Data nabycia", "Typ", "Ilosc", "Koszt EUR",
                      "Koszt PLN", "Przychod EUR", "Przychod PLN", "Kurs NBP lotu",
                      "Data kursu lotu", "Tabela NBP lotu", "Data sprzedazy",
                      "Kurs NBP sprzedazy", "Tabela NBP sprzedazy"])
    for row in trace_rows:
        ws_trace.append([
            row["lot_id"], row["acquired_date"], row["lot_type"], row["quantity"],
            row["cost_eur"], row["cost_pln"], row["revenue_eur"], row["revenue_pln"],
            row["lot_nbp_rate"], row["lot_nbp_rate_date"], row["lot_table_no"],
            row["sale_date"], row["sale_nbp_rate"], row["sale_table_no"]])

    ws_div = wb.create_sheet("Dywidendy")
    ws_div.append(["Rok", year])
    ws_div.append(["Liczba dywidend", report["section_g"]["dividend_count"]])
    ws_div.append(["Brutto PLN", report["section_g"]["gross_pln"]])
    ws_div.append(["Pobrane u zrodla PLN", report["section_g"]["withholding_paid_pln"]])
    ws_div.append(["Doplata w PL PLN", report["section_g"]["pl_tax_due_pln"]])
    ws_div.append(
        ["Do odzyskania z Vero PLN", report["section_g"]["reclaimable_from_finland_pln"]])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
